import io
from datetime import datetime
import pytz
import pyzipper
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from auth import get_current_user, require_export_privilege
from app.database import get_db
from app import models

router = APIRouter(prefix="/api/v1/analytics", tags=["Analytics Exports"])

@router.get("/export")
def export_analytics_report(db: Session = Depends(get_db), current_user = Depends(require_export_privilege)):
    try:
        is_global = current_user.role in ['SUPER_ADMIN', 'ADMIN', 'RPC'] or str(current_user.region).upper() in ['KMP HEADQUARTERS', 'POLICE HEADQUARTERS']
        
        # 1. ORM Models
        CrimeModel = getattr(models, 'Crime_Reports', getattr(models, 'CrimeReports', getattr(models, 'Reports', None)))
        StatsModel = getattr(models, 'Operational_Statistics', getattr(models, 'OperationalStatistics', getattr(models, 'Stats', None)))
        StoryModel = getattr(models, 'Success_Stories', getattr(models, 'SuccessStories', getattr(models, 'Stories', None)))
        NomModel = getattr(models, 'Nominal_Roll', getattr(models, 'NominalRoll', getattr(models, 'User', None)))
        AgricModel = getattr(models, 'Agricultural_Crime_Summary', None)

        def get_scoped_query(ModelClass):
            if not ModelClass:
                return []
            q = db.query(ModelClass)
            if not is_global and hasattr(ModelClass, 'region'):
                q = q.filter(ModelClass.region == current_user.region)
            return q.all()

        cr_records = get_scoped_query(CrimeModel)
        ops_records = get_scoped_query(StatsModel)
        ss_records = get_scoped_query(StoryModel)
        nom_records = get_scoped_query(NomModel)
        agric_records = get_scoped_query(AgricModel)

        # 2. Build Specialized Datasets

        # --- A. Agricultural Crimes Sub-Categories (Animals, Produce, Equipment) ---
        agric_breakdown = {"ANIMALS": [0, 0], "PRODUCE": [0, 0], "EQUIPMENT": [0, 0]}
        for ag in agric_records:
            rep_type = str(getattr(ag, 'agric_crime_report', '')).upper()
            stolen = getattr(ag, 'number_count', 0) or 0
            recovered = getattr(ag, 'recoveries', 0) or 0
            if "ANIMAL" in rep_type or "LIVESTOCK" in rep_type or "CATTLE" in rep_type:
                agric_breakdown["ANIMALS"][0] += stolen
                agric_breakdown["ANIMALS"][1] += recovered
            elif "EQUIPMENT" in rep_type or "IMPLEMENT" in rep_type or "MACHINE" in rep_type:
                agric_breakdown["EQUIPMENT"][0] += stolen
                agric_breakdown["EQUIPMENT"][1] += recovered
            else:
                agric_breakdown["PRODUCE"][0] += stolen
                agric_breakdown["PRODUCE"][1] += recovered

        agric_cat_data = [
            ["ANIMALS (Livestock & Wildlife)", agric_breakdown["ANIMALS"][0], agric_breakdown["ANIMALS"][1]],
            ["PRODUCE (Crops & Harvest)", agric_breakdown["PRODUCE"][0], agric_breakdown["PRODUCE"][1]],
            ["EQUIPMENT (Farm Implements & Tools)", agric_breakdown["EQUIPMENT"][0], agric_breakdown["EQUIPMENT"][1]]
        ]

        # --- B. Manpower Analysis (Corrected NCOs including HCM and HC with M/F split) ---
        officer_ranks = ['CP', 'ACP', 'SSP', 'SP', 'SASP', 'ASP', 'IP', 'AIP']
        nco_ranks = ['HCM', 'HC', 'S/SGT', 'SGT', 'CPL', 'L/CPL', 'PC', 'PPC', 'SPC']
        all_ranks = officer_ranks + nco_ranks

        manpower_matrix = {}
        total_male_general = 0
        total_female_general = 0

        for n in nom_records:
            reg = str(getattr(n, 'region', 'KMP HEADQUARTERS') or 'KMP HEADQUARTERS').upper()
            stat = str(getattr(n, 'station', 'HQ') or 'HQ').upper()
            rnk = str(getattr(n, 'rank', 'PC') or 'PC').upper()
            sex = str(getattr(n, 'sex', 'M') or 'M').upper()

            if reg not in manpower_matrix:
                manpower_matrix[reg] = {}
            if stat not in manpower_matrix[reg]:
                manpower_matrix[reg][stat] = {rk: {'M': 0, 'F': 0} for rk in all_ranks}

            if rnk in manpower_matrix[reg][stat]:
                if 'F' in sex:
                    manpower_matrix[reg][stat][rnk]['F'] += 1
                    total_female_general += 1
                else:
                    manpower_matrix[reg][stat][rnk]['M'] += 1
                    total_male_general += 1

        manpower_table_rows = []
        for reg, stations in manpower_matrix.items():
            manpower_table_rows.append([f"REGION: {reg}", "", "", ""] + [""] * (len(all_ranks) * 2))
            for stat, ranks_data in stations.items():
                stat_m = sum(v['M'] for v in ranks_data.values())
                stat_f = sum(v['F'] for v in ranks_data.values())
                stat_total = stat_m + stat_f
                row_entry = [reg, stat, stat_total, stat_m, stat_f]
                for rk in all_ranks:
                    row_entry.append(ranks_data[rk]['M'])
                    row_entry.append(ranks_data[rk]['F'])
                manpower_table_rows.append(row_entry)

        # --- C. Success Stories (One-Line Bullet Sentence Format) ---
        success_data = []
        for st in ss_records:
            date_val = str(getattr(st, 'date', ''))
            reg_val = getattr(st, 'region', '')
            stat_val = getattr(st, 'station', '')
            narrative = getattr(st, 'narrative', getattr(st, 'title', 'Successful operation executed.'))
            bullet_sentence = f"• Successful operational breakthrough achieved on {date_val} at {stat_val} ({reg_val}): {narrative}."
            success_data.append([reg_val, stat_val, bullet_sentence])

        # --- D. Disruptive Ops Grouped Weekly per Station within Regional Blocks ---
        disruptive_data = []
        region_ops_totals = {}
        for s in ops_records:
            reg = getattr(s, 'region', 'GENERAL')
            stat = getattr(s, 'station', 'N/A')
            wk = getattr(s, 'date', 'WEEKLY PERIOD')
            
            if reg not in region_ops_totals:
                region_ops_totals[reg] = {"arrested": 0, "bond": 0, "caution": 0, "pending": 0, "court": 0, "released": 0, "remanded": 0, "convicted": 0}
            
            arr = getattr(s, 'arrested', 0) or 0
            bon = getattr(s, 'given_bond', 0) or 0
            cau = getattr(s, 'cautioned', 0) or 0
            pen = getattr(s, 'pending_court', 0) or 0
            tak = getattr(s, 'taken_to_court', 0) or 0
            rel = getattr(s, 'released', 0) or 0
            rem = getattr(s, 'remanded', 0) or 0
            con = getattr(s, 'convicted', 0) or 0

            region_ops_totals[reg]["arrested"] += arr
            region_ops_totals[reg]["bond"] += bon
            region_ops_totals[reg]["caution"] += cau
            region_ops_totals[reg]["pending"] += pen
            region_ops_totals[reg]["court"] += tak
            region_ops_totals[reg]["released"] += rel
            region_ops_totals[reg]["remanded"] += rem
            region_ops_totals[reg]["convicted"] += con

            disruptive_data.append([str(wk), reg, stat, arr, bon, cau, pen, tak, rel, rem, con])

        # --- E. Comparative Distribution & Volume ---
        comp_counts = {}
        for r in cr_records:
            cat = getattr(r, 'offence', 'GENERAL CRIME') or 'GENERAL CRIME'
            comp_counts[cat] = comp_counts.get(cat, 0) + 1
        comp_data = [[k, v] for k, v in sorted(comp_counts.items(), key=lambda x: x[1], reverse=True)]

        # --- F. Master Summary Table ---
        summary_table_data = [
            ["Total General Manpower (Force-Wide)", len(nom_records)],
            ["Total General Male Personnel", total_male_general],
            ["Total General Female Personnel", total_female_general],
            ["Total Success Stories (General Force-Wide)", len(ss_records)],
            ["Total Recorded Incidents / Crime Reports", len(cr_records)],
            ["Total Disruptive Operations Logs", len(ops_records)],
            ["Total Suspects Arrested Force-Wide", sum(getattr(s, 'arrested', 0) or 0 for s in ops_records)],
            ["Total Convictions Obtained Force-Wide", sum(getattr(s, 'convicted', 0) or 0 for s in ops_records)]
        ]
        
        for r_name, totals in region_ops_totals.items():
            summary_table_data.append([f"Disruptive Ops Total - Region: {r_name} (Arrested / Convicted)", f"Arrested: {totals['arrested']} | Convicted: {totals['convicted']}"])

        # 3. Build Excel Workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        section_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        section_font = Font(color="FFFFFF", bold=True, size=11)

        def add_individual_sheet(title, headers, rows):
            ws = wb.create_sheet(title=title)
            ws.append(["SN"] + headers)
            for cell in ws[1]:
                cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center")
            for idx, r in enumerate(rows, 1):
                ws.append([idx] + list(r))
            for col in ws.columns:
                max_len = max([len(str(cell.value or '')) for cell in col], default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

        manpower_headers = ["Region", "Station", "Total", "Male", "Female"] + [item for r in all_ranks for item in (f"{r} (M)", f"{r} (F)")]
        add_individual_sheet("Manpower Analysis", manpower_headers, manpower_table_rows)
        add_individual_sheet("Agricultural Crimes", ["Sub-Category", "Stolen Count", "Recovered Count"], agric_cat_data)
        add_individual_sheet("Success Stories", ["Region", "Station", "Operational Success Highlight (One-Line Bullet)"], success_data)
        add_individual_sheet("Disruptive Ops", ["Weekly Period", "Region", "Station", "Arrested", "Bonded", "Cautioned", "Pending Court", "To Court", "Released", "Remanded", "Convicted"], disruptive_data)
        add_individual_sheet("Comparative Trends", ["Category / Offence", "Total Volume"], comp_data)
        add_individual_sheet("Master Summary Aggregates", ["Operational Metric Attribute", "Aggregate Value / Total"], summary_table_data)

        # 4. Build Master 'General Analytics' Sheet
        ws_gen = wb.create_sheet(title="General Analytics", index=0)
        
        def append_stacked_section(section_title, headers, rows):
            ws_gen.append([section_title.upper()])
            title_cell = ws_gen.cell(row=ws_gen.max_row, column=1)
            title_cell.fill = section_fill
            title_cell.font = section_font
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            ws_gen.append(["SN"] + headers)
            header_row_idx = ws_gen.max_row
            for col_idx in range(1, len(headers) + 2):
                c = ws_gen.cell(row=header_row_idx, column=col_idx)
                c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal="center", vertical="center")

            if not rows:
                ws_gen.append(["—", "No records captured for this analytical attribute."])
            else:
                for idx, r in enumerate(rows, 1):
                    ws_gen.append([idx] + list(r))
            ws_gen.append([])

        append_stacked_section("1. Manpower Analysis (Officers & NCOs breakdown with HCM/HC)", manpower_headers, manpower_table_rows)
        append_stacked_section("2. Agricultural Crimes Breakdown (Animals, Produce, Equipment)", ["Sub-Category", "Stolen Count", "Recovered Count"], agric_cat_data)
        append_stacked_section("3. Success Stories & Breakthroughs (One-Line Bullet Sentences)", ["Region", "Station", "Operational Success Highlight"], success_data)
        append_stacked_section("4. Disruptive Operations Grouped Weekly by Station", ["Weekly Period", "Region", "Station", "Arrested", "Bonded", "Cautioned", "Pending Court", "To Court", "Released", "Remanded", "Convicted"], disruptive_data)
        append_stacked_section("5. Comparative Distribution & Volume Trends", ["Category / Offence", "Total Volume"], comp_data)
        append_stacked_section("6. Master Summary Table (General & Regional Totals)", ["Operational Metric Attribute", "Aggregate Value / Total"], summary_table_data)

        for col in ws_gen.columns:
            max_len = max([len(str(cell.value or '')) for cell in col], default=0)
            ws_gen.column_dimensions[col[0].column_letter].width = min(max_len + 3, 55)

        # 5. Encrypt into AES-256 ZIP Archive
        excel_stream = io.BytesIO()
        wb.save(excel_stream)

        zip_stream = io.BytesIO()
        eat_time = datetime.now(pytz.timezone("Africa/Nairobi")).replace(tzinfo=None)
        fnum_clean = str(current_user.fnum).replace('/', '_').upper()
        zip_password = str(current_user.fnum).strip().encode('utf-8')

        with pyzipper.AESZipFile(zip_stream, 'w', compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES) as zf:
            zf.setpassword(zip_password)
            zf.writestr(f"{fnum_clean}_Analytics_Report_{eat_time.strftime('%Y%m%d')}.xlsx", excel_stream.getvalue())

        zip_stream.seek(0)
        return StreamingResponse(
            zip_stream,
            media_type="application/zip",
            headers={
                'Content-Disposition': f'attachment; filename="SECURE_ANALYTICS_REPORT_{eat_time.strftime("%Y%m%d")}.zip"',
                'Access-Control-Expose-Headers': 'Content-Disposition'
            }
        )
    except Exception as e:
        print(f"Analytics Export Error: {e}")
        raise HTTPException(status_code=500, detail=f"Analytics export compilation failed: {str(e)}")