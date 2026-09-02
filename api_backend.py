import io
import os
import gc
import re
import html
import uuid
import asyncio
import secrets  
import string    
import json
import base64
from datetime import datetime, timedelta
from typing import Optional, List, Union
from docx.shared import Pt
from docx.enum.text import WD_ALIGN_PARAGRAPH
import urllib.parse
import pymupdf
import zipfile

import pytz
import uvicorn
import pyzipper
import pandas as pd
import numpy as np

import boto3
from botocore.exceptions import ClientError
from dotenv import load_dotenv

from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Form, BackgroundTasks, Request
from fastapi.responses import StreamingResponse, FileResponse, RedirectResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from fastapi_mail import ConnectionConfig, FastMail, MessageSchema
from jose import jwt, JWTError
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from apscheduler.schedulers.background import BackgroundScheduler
from sqlalchemy import Column, Integer, text
from docx.shared import Pt, RGBColor
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from pptx import Presentation
from pptx.util import Inches, Pt as PPTXPt
from pptx.dml.color import RGBColor as PPTXRGBColor
from sqlalchemy.orm.attributes import flag_modified
from passlib.context import CryptContext

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

from urllib.parse import unquote
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

# Internal Imports
from app import models, schemas, database
from app.database import engine, get_db, get_logs_db, SQLALCHEMY_DATABASE_URL
from app.core import security
from auth import router as auth_router
from docx import Document
from app.schemas import AgricStatsCreate, AgricStatsResponse

# ==========================================
# 0. LOAD ENVIRONMENT VARIABLES & CONFIG
# ==========================================
load_dotenv()

app = FastAPI(title="KMP Centralised Security Data Management System")

from routers import (
    document_upload,
    general_documents,
    command_templates,
    agric_summary,
    nominal_roll,
    crime_registry,
    lockup_matrix,
    establishments,
    success_stories,
    admin_communication,
    ai_router,
    analytics_export,  # Import analytics_export correctly
)

# Include routers once
app.include_router(document_upload.router)
app.include_router(general_documents.router)
app.include_router(command_templates.router)
app.include_router(agric_summary.router)
app.include_router(nominal_roll.router)
app.include_router(crime_registry.router)
app.include_router(lockup_matrix.router)
app.include_router(establishments.router)
app.include_router(success_stories.router)
app.include_router(analytics_export.router)
app.include_router(admin_communication.router)
app.include_router(auth_router, prefix="/api/auth")
app.include_router(ai_router.router)
app.include_router(auth_router)
app.include_router(auth_router, prefix="/api/auth")
app.include_router(auth_router, prefix="/api/v1/auth")
app.include_router(auth_router, prefix="/api/v1/users")

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    print(f"Internal Command Error Traceback: {str(exc)}")
    return JSONResponse(
        status_code=500,
        content={"detail": "An internal command processing error occurred. Please try again later."},
    )

limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

conf = ConnectionConfig(
    MAIL_USERNAME=os.getenv("MAIL_USERNAME"),
    MAIL_PASSWORD=os.getenv("MAIL_PASSWORD"),
    MAIL_FROM=os.getenv("MAIL_FROM"),
    MAIL_PORT=587,
    MAIL_SERVER="smtp.gmail.com",
    MAIL_STARTTLS=True,
    MAIL_SSL_TLS=False
)

s3_client = boto3.client(
    "s3",
    aws_access_key_id=os.getenv("AWS_ACCESS_KEY_ID"),
    aws_secret_access_key=os.getenv("AWS_SECRET_ACCESS_KEY"),
    region_name=os.getenv("AWS_REGION")
)
BUCKET_NAME = os.getenv("AWS_BUCKET_NAME")

# ==========================================
# UNIVERSAL EXPORT HELPERS
# ==========================================
def clean_html_for_export(raw_text):
    """Strips HTML tags and converts entities into readable plain text for Excel exports."""
    if not raw_text or not isinstance(raw_text, str):
        return raw_text
    
    text = html.unescape(raw_text)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def get_officer_signature(user):
    if not user:
        return "UNKNOWN COMMANDER"
    fnum = (user.fnum or "").strip()
    rank = (user.rank or "").strip()
    name = (user.name or "").strip()
    return f"{fnum} {rank} {name}".strip().upper()

def sanitize_df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    for col in df.select_dtypes(include=['datetimetz', 'datetime', 'datetime64[ns, UTC]', 'datetime64[ns]']).columns:
        try:
            df[col] = df[col].dt.tz_localize(None)
        except Exception:
            df[col] = df[col].astype(str)
    for col in df.select_dtypes(include=['object']).columns:
        df[col] = df[col].apply(lambda x: x.replace(tzinfo=None) if isinstance(x, datetime) and x.tzinfo is not None else x)
        # Apply HTML scrubber to text fields during export
        if col in ['narrative', 'comment', 'message', 'details', 'archive_reason']:
            df[col] = df[col].apply(clean_html_for_export)
    return df

def serialize_model_row(row):
    if not row:
        return {}
    if isinstance(row, dict):
        d = row.copy()
    elif hasattr(row, '__dict__'):
        d = row.__dict__.copy()
        d.pop('_sa_instance_state', None)
    else:
        return {}
    for k, v in d.items():
        if isinstance(v, datetime):
            d[k] = v.strftime("%Y-%m-%d %H:%M:%S")
        elif hasattr(v, 'isoformat'):
            d[k] = v.isoformat()
        elif hasattr(v, '__float__') and not isinstance(v, (int, float, str, bool)):
            d[k] = float(v)
    return d

# ==========================================
# 2. STARTUP & MIDDLEWARE
# ==========================================
@app.get("/")
@app.head("/")
def health_check():
    return {
        "status": "online", 
        "message": "KMP Centralised Security API is running securely."
    }

try:
    models.Base.metadata.create_all(bind=engine, checkfirst=True)
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS last_active_at TIMESTAMP"))
except Exception as e:
    print(f"Startup metadata notice: {e}")

os.makedirs("uploads", exist_ok=True)
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173", 
        "http://127.0.0.1:5173",
        "https://kmp-tracker-system-centralised-secu.vercel.app",
        "https://kmp-tracker-system-centralised-security-management-adj4h23x4.vercel.app",
        "https://kmp-tracker-system-centralised-security-management-od0odfzxy.vercel.app"
    ],
    allow_origin_regex=r"https://kmp-tracker.*\.vercel\.app",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)

# ==========================================
# 3. SECURITY & DEPENDENCIES (RAM-ONLY SECURE)
# ==========================================
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login", auto_error=False)

def get_eat_time():
    eat_tz = pytz.timezone('Africa/Nairobi')
    return datetime.now(eat_tz).strftime('%Y-%m-%d %H:%M:%S')

def get_current_user(request: Request, db: Session = Depends(get_db)):
    token = None
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]
    if not token:
        token = request.cookies.get("access_token") or request.cookies.get("kmp_authToken")

    if not token:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Not authenticated")

    try:
        payload = jwt.decode(token, security.SECRET_KEY, algorithms=[security.ALGORITHM])
        fnum: str = payload.get("sub")
        if fnum is None:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Session expired")
    
    clean_fnum = str(fnum).strip().upper()
    user = db.query(models.Users).filter(
        func.trim(func.upper(models.Users.fnum)) == clean_fnum
    ).first()
    
    if user is None:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found")

    try:
        eat_tz = pytz.timezone('Africa/Nairobi')
        user.last_active_at = datetime.now(eat_tz).replace(tzinfo=None)
        db.commit()
    except Exception:
        db.rollback()

    if user.role != "SUPER_ADMIN":
        config_check = db.query(models.SystemConfig).filter(
            models.SystemConfig.config_key == "peer_delegation_active"
        ).first()
        if config_check and str(config_check.config_value).strip().upper() == "FALSE":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Application Lockdown: System access restricted."
            )

    return user

def require_admin(current_user: models.Users = Depends(get_current_user)):
    user_role = str(current_user.role).strip().upper() if current_user.role else ""
    if "ADMIN" not in user_role and "RPC" not in user_role:
        raise HTTPException(status_code=403, detail="Clearance Denied: Admin privileges required.")
    return current_user

def require_export_privilege(current_user: models.Users = Depends(get_current_user)):
    user_role = str(current_user.role).strip().upper() if current_user.role else ""
    perms = current_user.permissions or {}
    if user_role not in ["ADMIN", "SUPER_ADMIN", "RPC"] and not perms.get("export_data", False):
        raise HTTPException(status_code=403, detail="Clearance Denied: Data Export Privileges Required.")
    return current_user

def log_semantic_audit(db, fnum: str, action: str, target_identifier: str, changes: dict, remarks: str = ""):
    try:
        formatted_details = f"Target: {target_identifier} | Changes: " + ", ".join(
            [f"{k}: {v[0]} -> {v[1]}" for k, v in changes.items()]
        ) + f" | Remarks: {remarks}"
        
        audit_model = getattr(models, 'Audit_Logs', getattr(models, 'AuditLogs', None))
        if audit_model:
            new_audit = audit_model(
                event_type=action,
                target_user=target_identifier,
                status="SUCCESS",
                details=formatted_details,
                user_fnum=fnum,
                created_at=get_eat_time()
            )
            db.add(new_audit)
            db.commit()
    except Exception as e:
        print(f"Audit Log Failed: {e}")
        db.rollback()

# ==========================================
# 4. USERS, HEARTBEAT & ACTIVITY LOGS
# ==========================================
@app.get("/api/v1/users")
def get_all_active_users(db: Session = Depends(get_db), current_user: models.Users = Depends(get_current_user)):
    try:
        users = db.query(models.Users).filter(models.Users.is_approved == True).all()
        return [
            {
                "fnum": u.fnum, "name": u.name, "rank": u.rank, "role": u.role, 
                "station": u.station, "region": u.region, "division": u.division,
                "position": u.position, "email": u.email, "phone": u.phone,
                "ipps": u.ipps, "sex": u.sex, "profile_photo_path": u.profile_photo_path,
                "permissions": u.permissions
            } for u in users
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/v1/general-documents")
def get_general_documents(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    try:
        DocModel = getattr(models, 'GeneralDocuments', getattr(models, 'General_Documents', None))
        if not DocModel:
            return []
        docs = db.query(DocModel).order_by(DocModel.id.desc()).all()
        return [serialize_model_row(d) for d in docs]
    except Exception as e:
        print(f"General Documents Fetch Error: {e}")
        return []

@app.get("/api/v1/templates/list")
def get_command_templates(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    try:
        TemplateModel = getattr(models, 'CommandTemplate', getattr(models, 'command_templates', None))
        if not TemplateModel:
            return []
        templates = db.query(TemplateModel).order_by(TemplateModel.id.desc()).all()
        results = []
        for t in templates:
            filename = getattr(t, 'file_name', None) or getattr(t, 's3_url', '') or "document"
            ext = filename.split('.')[-1].lower() if '.' in filename else "unknown"
            if ext in ['docx', 'doc']: file_type = "Word Document"
            elif ext in ['xlsx', 'xls']: file_type = "Excel Spreadsheet"
            elif ext in ['pptx', 'ppt']: file_type = "PowerPoint Presentation"
            elif ext == 'pdf': file_type = "PDF Document"
            else: file_type = "Command Template"

            last_up = getattr(t, 'upload_date', getattr(t, 'created_at', None))
            date_str = last_up.strftime("%Y-%m-%d") if isinstance(last_up, datetime) else str(last_up or "")

            results.append({
                "id": getattr(t, 'id', 1),
                "name": filename,
                "type": file_type,
                "date": date_str,
                "size": getattr(t, 'file_size', "N/A"),
                "file_path": getattr(t, 'file_path', getattr(t, 's3_url', '')),
                "region": getattr(t, 'region', "KMP HEADQUARTERS"),
                "station": getattr(t, 'station', "HQ")
            })
        return results
    except Exception as e:
        print(f"Templates List Error Traceback: {str(e)}")
        return []

@app.get("/api/v1/weekly-reports/list")
def get_weekly_reports_list(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    try:
        ReportModel = getattr(models, 'WeeklyReports', getattr(models, 'weekly_reports', getattr(models, 'Reports', None)))
        if not ReportModel:
            return []
        reports = db.query(ReportModel).order_by(ReportModel.id.desc()).all()
        return [serialize_model_row(r) for r in reports]
    except Exception as e:
        print(f"Weekly Reports Fetch Error: {e}")
        return []

@app.get("/api/v1/admin/pending-users")
def get_pending_users(db: Session = Depends(get_db), current_user: models.Users = Depends(get_current_user)):
    user_role = str(current_user.role).strip().upper() if current_user.role else ""
    if "ADMIN" not in user_role and "RPC" not in user_role:
        raise HTTPException(status_code=403, detail="Unauthorized access.")
    try:
        pending = db.query(models.Users).filter(models.Users.is_approved == False).all()
        return [
            {
                "id": getattr(u, 'id', 0),
                "fnum": u.fnum, "name": u.name, "rank": u.rank,
                "station": u.station, "region": u.region, "role": u.role,
                "email": u.email, "phone": u.phone, "ipps": u.ipps, "sex": u.sex
            } for u in pending
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/v1/requests")
def get_system_requests(db: Session = Depends(get_db), current_user: models.Users = Depends(get_current_user)):
    user_role = str(current_user.role).strip().upper() if current_user.role else ""
    if "ADMIN" not in user_role and "RPC" not in user_role:
        raise HTTPException(status_code=403, detail="Unauthorized access.")
    try:
        requests = db.query(models.Users).filter(models.Users.is_approved == False).all()
        return [
            {
                "id": getattr(r, 'id', 0),
                "fnum": r.fnum, "name": r.name, "rank": r.rank,
                "station": r.station, "region": r.region, "role": r.role
            } for r in requests
        ]
    except Exception as e:
        return []

@app.get("/api/v1/audit-logs")
def get_audit_logs(db: Session = Depends(get_db), current_user: models.Users = Depends(get_current_user)):
    user_role = str(current_user.role).strip().upper() if current_user.role else ""
    if "ADMIN" not in user_role and "RPC" not in user_role:
        raise HTTPException(status_code=403, detail="Clearance Denied: Admin privileges required.")
    try:
        AuditModel = getattr(models, 'Audit_Logs', getattr(models, 'AuditLogs', None))
        if not AuditModel:
            return []
        logs = db.query(AuditModel).order_by(AuditModel.id.desc()).limit(100).all()
        return [
            {
                "id": log.id,
                "event_type": getattr(log, 'event_type', 'ACTION'),
                "target_user": getattr(log, 'target_user', ''),
                "status": getattr(log, 'status', 'SUCCESS'),
                "details": getattr(log, 'details', ''),
                "user_fnum": getattr(log, 'user_fnum', ''),
                "created_at": str(getattr(log, 'created_at', ''))
            } for log in logs
        ]
    except Exception as e:
        print(f"Audit log fetch error: {e}")
        return []

@app.put("/api/v1/users/{fnum:path}/access")
@app.post("/api/v1/admin/bulk-permissions")
@app.put("/api/v1/admin/bulk-permissions")
def update_user_access(
    data: dict, 
    fnum: Optional[str] = None, 
    db: Session = Depends(get_db), 
    current_user: models.Users = Depends(require_admin)
):
    target_fnum = fnum or data.get("fnum") or data.get("user_fnum")
    if not target_fnum:
        raise HTTPException(status_code=400, detail="Officer force number (fnum) is required.")

    clean_fnum = unquote(unquote(target_fnum)).strip().upper()
    
    user = db.query(models.Users).filter(
        func.trim(func.upper(models.Users.fnum)) == clean_fnum
    ).first()
    
    if not user:
        alt_fnum = clean_fnum.replace('/', '')
        user = db.query(models.Users).filter(
            func.trim(func.upper(models.Users.fnum)) == alt_fnum
        ).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="Officer record not found.")

    if "role" in data and data["role"]:
        user.role = str(data["role"]).strip().upper()

    # 🟢 AUTOMATICALLY SET APPROVAL TO TRUE UPON ACCESS ASSIGNMENT
    user.is_approved = True

    if "is_approved" in data:
        user.is_approved = bool(data["is_approved"])

    if "permissions" in data and data["permissions"] is not None:
        merged_perms = dict(user.permissions or {})
        if isinstance(data["permissions"], dict):
            merged_perms.update(data["permissions"])
        user.permissions = merged_perms
        flag_modified(user, "permissions")

    try:
        db.commit()
        db.refresh(user)
        return {
            "status": "success",
            "message": f"Access matrix and approval permanently committed for {user.fnum}",
            "permissions": user.permissions,
            "role": user.role,
            "is_approved": user.is_approved
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database commit error: {str(e)}")

@app.get("/api/v1/admin/reset-requests")
def get_reset_requests(db: Session = Depends(get_db), current_user: models.Users = Depends(get_current_user)):
    try:
        if current_user.role not in ["ADMIN", "SUPER_ADMIN", "RPC"]:
            raise HTTPException(status_code=403, detail="Unauthorized access.")
        target_model = getattr(models, 'Password_Reset_Requests', getattr(models, 'PasswordResetRequests', None))
        if target_model:
            resets = db.query(target_model).all()
            return [{"id": r.id, "fnum": r.fnum, "status": r.status} for r in resets]
        return []
    except Exception as e:
        return []

@app.post("/api/v1/admin/execute-reset/{req_id}")
def execute_password_reset(
    req_id: int,
    action: str = Form(...),
    db: Session = Depends(get_db),
    current_user: models.Users = Depends(get_current_user)
):
    if current_user.role not in ["ADMIN", "SUPER_ADMIN", "RPC"]:
        raise HTTPException(status_code=403, detail="Unauthorized access.")

    TargetModel = getattr(models, 'Password_Reset_Requests', getattr(models, 'PasswordResetRequests', None))
    if not TargetModel:
        raise HTTPException(status_code=404, detail="Password reset model not initialized.")

    reset_req = db.query(TargetModel).filter(TargetModel.id == req_id).first()
    if not reset_req:
        raise HTTPException(status_code=404, detail="Password reset request not found.")

    if action.upper() == "APPROVE":
        temp_password = "UPF" + secrets.token_hex(3).upper()
        hashed_pw = pwd_context.hash(temp_password)
        target_user = db.query(models.Users).filter(
            func.trim(func.upper(models.Users.fnum)) == str(reset_req.fnum).strip().upper()
        ).first()

        if target_user:
            target_user.hashed_password = hashed_pw

        db.delete(reset_req)
        db.commit()

        return {"status": "success", "new_password": temp_password}
    else:
        db.delete(reset_req)
        db.commit()
        return {"status": "success", "message": "Password reset request rejected."}

class HeartbeatPayload(BaseModel):
    fnum: Optional[str] = None

@app.post("/api/v1/users/heartbeat")
@app.post("/api/v1/users/heartbeat/")
def heartbeat(
    payload: Optional[HeartbeatPayload] = None,
    token: Optional[str] = Depends(OAuth2PasswordBearer(tokenUrl="/api/auth/login", auto_error=False)),
    db: Session = Depends(get_db)
):
    user = None
    if token:
        try:
            jwt_data = jwt.decode(token, security.SECRET_KEY, algorithms=[security.ALGORITHM])
            token_fnum = jwt_data.get("sub")
            if token_fnum:
                user = db.query(models.Users).filter(
                    func.trim(func.upper(models.Users.fnum)) == str(token_fnum).strip().upper()
                ).first()
        except JWTError:
            pass

    if not user and payload and payload.fnum:
        user = db.query(models.Users).filter(
            func.trim(func.upper(models.Users.fnum)) == str(payload.fnum).strip().upper()
        ).first()

    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid session credentials")

    try:
        eat_tz = pytz.timezone('Africa/Nairobi')
        user.last_active_at = datetime.now(eat_tz).replace(tzinfo=None)
        db.commit()
        
        expire = datetime.utcnow() + timedelta(minutes=60)
        new_token = jwt.encode(
            {"sub": str(user.fnum).strip().upper(), "exp": expire},
            security.SECRET_KEY,
            algorithm=security.ALGORITHM
        )
        
        return {"status": "alive", "user": user.fnum, "new_token": new_token}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/v1/users/online")
def get_online_users(db: Session = Depends(get_db), current_user: models.Users = Depends(get_current_user)):
    eat_tz = pytz.timezone('Africa/Nairobi')
    now_eat = datetime.now(eat_tz).replace(tzinfo=None)
    threshold = now_eat - timedelta(minutes=2)
    
    active_users = db.query(models.Users).filter(
        models.Users.is_approved == True,
        models.Users.last_active_at >= threshold
    ).all()
    
    return [
        {
            "fnum": u.fnum,
            "name": u.name,
            "station": u.station,
            "profile_photo_path": u.profile_photo_path
        } for u in active_users
    ]

@app.get("/api/v1/activity-logs")
def get_system_activity_logs(db: Session = Depends(get_logs_db), current_user: models.Users = Depends(get_current_user)):
    try:
        if current_user.role not in ["ADMIN", "SUPER_ADMIN", "RPC"]:
            raise HTTPException(status_code=403, detail="Unauthorized access.")
        
        logs = db.query(models.Activity_Logs).order_by(models.Activity_Logs.id.desc()).limit(100).all()
        return [
            {
                "id": log.id,
                "created_at": log.created_at.isoformat() if hasattr(log.created_at, 'isoformat') else str(log.created_at),
                "fnum": log.fnum or '',
                "action": log.action or '',
                "module": log.module or '',
                "details": log.details or ''
            } for log in logs
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch activity logs: {str(e)}")

@app.post("/api/v1/activity-logs")
def create_system_activity_log(data: dict, db: Session = Depends(get_logs_db), current_user: models.Users = Depends(get_current_user)):
    try:
        page = data.get("module", data.get("page_accessed", "UNKNOWN"))
        act = data.get("action", "PAGE_ACCESS")
        details = data.get("details", f"Officer {current_user.name} ({current_user.fnum}) accessed {page}")
        
        new_activity = models.Activity_Logs(
            fnum=current_user.fnum,
            action=act,
            module=page,
            details=details,
            created_at=get_eat_time()
        )
        db.add(new_activity)
        db.commit()
        return {"status": "success"}
    except Exception as e:
        db.rollback()
        return {"status": "error", "detail": str(e)}

@app.post("/api/v1/system/log-session")
def log_user_session(data: dict, db: Session = Depends(get_db)):
    fnum = data.get("fnum")
    if hasattr(models, 'Audit_Logs') and fnum:
        log_semantic_audit(
            db=db, fnum=fnum, action="OFFICER_AUTHENTICATION", 
            target_identifier="SYSTEM", changes={}, remarks="Secure session initiated via Dashboard Gateway"
        )
    return {"status": "success"}

@app.get("/api/v1/users/recipients-list")
@app.get("/api/v1/communications/recipients-list")
def get_recipients_list(db: Session = Depends(get_db), current_user: models.Users = Depends(get_current_user)):
    try:
        users = db.query(models.Users).filter(
            models.Users.role != 'REVOKED',
            models.Users.is_approved == True
        ).all()
        return [
            {
                "fnum": u.fnum,
                "name": u.name,
                "rank": u.rank,
                "region": u.region,
                "station": u.station,
                "position": u.position,
                "role": u.role
            } for u in users
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch recipients: {str(e)}")

@app.get("/api/v1/reports/establishments-json")
def get_establishments_json(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    try:
        est_model = getattr(models, 'Establishments', getattr(models, 'establishments', None))
        nom_model = getattr(models, 'Nominal_Roll', getattr(models, 'NominalRoll', None))
        
        est = db.query(est_model).all() if est_model else []
        nom = db.query(nom_model).all() if nom_model else []
        
        return {
            "establishments": [serialize_model_row(e) for e in est],
            "nominal_rolls": [serialize_model_row(n) for n in nom]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to compile HR Ledger: {str(e)}")

@app.get("/api/v1/reports/consolidated-ledger")
def get_consolidated_ledger(
    start_date: Optional[str] = None, 
    end_date: Optional[str] = None, 
    db: Session = Depends(get_db), 
    current_user = Depends(get_current_user)
):
    try:
        CrimeModel = getattr(models, 'Crime_Reports', getattr(models, 'CrimeReports', None))
        StatsModel = getattr(models, 'Operational_Statistics', getattr(models, 'OperationalStatistics', None))
        StoryModel = getattr(models, 'Success_Stories', getattr(models, 'SuccessStories', None))
        EstModel = getattr(models, 'Establishments', getattr(models, 'establishments', None))
        NomModel = getattr(models, 'Nominal_Roll', getattr(models, 'NominalRoll', None))

        crimes = db.query(CrimeModel).all() if CrimeModel else []
        stats = db.query(StatsModel).all() if StatsModel else []
        stories = db.query(StoryModel).all() if StoryModel else []
        establishments = db.query(EstModel).all() if EstModel else []
        nominal_roll = db.query(NomModel).all() if NomModel else []
        
        return {
            "status": "success",
            "crimes": [serialize_model_row(c) for c in crimes],
            "statistics": [serialize_model_row(s) for s in stats],
            "stories": [serialize_model_row(st) for st in stories],
            "establishments": [serialize_model_row(e) for e in establishments],
            "nominal_rolls": [serialize_model_row(n) for n in nominal_roll]
        }
    except Exception as e:
        print(f"Consolidated Ledger DB Query Error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch consolidated data: {str(e)}")

# ====================================================================
# 6. FULLY DYNAMIC INTELLIGENCE & EVENT-DRIVEN SCHEDULER
# ====================================================================
def run_weekly_tactical_briefing_job():
    eat_tz = pytz.timezone('Africa/Nairobi')
    now_eat = datetime.now(eat_tz).replace(tzinfo=None)
    one_week_ago = now_eat - timedelta(days=7)
    
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    db = SessionLocal()
    
    try:
        active_users = db.query(models.Users).filter(
            models.Users.is_approved == True,
            models.Users.email != None,
            models.Users.email != ""
        ).all()
        
        fm = FastMail(conf)
        
        async def process_and_send_emails():
            for user in active_users:
                station = user.station
                region = user.region
                is_global = user.role in ['SUPER_ADMIN', 'ADMIN', 'RPC'] or str(region).upper() in ['KMP HEADQUARTERS', 'POLICE HEADQUARTERS']
                
                crime_filter = "" if is_global else f" AND station = '{station}'"
                stats_filter = "" if is_global else f" AND station = '{station}'"
                
                crimes = db.execute(text(f"SELECT offence, narrative, status FROM reports WHERE created_at >= :start {crime_filter}"), {"start": one_week_ago}).fetchall()
                ops_stats = db.execute(text(f"SELECT arrests, given_bond, cautioned, remanded, convicted FROM stats WHERE date >= :start {stats_filter}"), {"start": one_week_ago.date()}).fetchall()
                
                total_arrests = sum(row.arrests or 0 for row in ops_stats)
                all_text = " ".join([f"{r.offence} {r.narrative}" for r in crimes]).upper()
                
                has_robbery = "ROBBERY" in all_text or "GUN" in all_text
                has_fire = "FIRE" in all_text or "ARSON" in all_text
                has_accident = "ACCIDENT" in all_text or "FATAL" in all_text
                has_murder = "MURDER" in all_text or "HOMICIDE" in all_text
                
                custom_actions = []
                if has_murder or has_robbery:
                    custom_actions.append("🔴 <b>High-Priority Security Spike:</b> Violent crime indicators (Robbery/Homicide) identified in weekly entries.")
                if has_fire:
                    custom_actions.append("🔥 <b>Public Safety Alert:</b> Fire or arson events logged.")
                if has_accident:
                    custom_actions.append("🚗 <b>Traffic Hazard Notice:</b> Traffic incidents/fatalities registered.")
                if total_arrests > 0:
                    custom_actions.append(f"⚖️ <b>Case Management:</b> {total_arrests} total arrests logged this week.")
                if not custom_actions:
                    custom_actions.append("✅ Operations stable for the period.")

                html_body = f"""
                <div style='font-family: Arial, sans-serif; color: #1e293b; max-w-[600px];'>
                    <h2 style='color: #0f172a; border-bottom: 2px solid #cbd5e1; padding-bottom: 10px;'>KMP Tactical Intelligence Briefing</h2>
                    <p><strong>Jurisdiction:</strong> {station} ({region})</p>
                    <p><strong>Officer:</strong> {user.rank} {user.name} ({user.fnum})</p>
                    <p>Below is your automated situational report for the past 7 days:</p>
                    <ul>
                        {''.join([f"<li style='margin-bottom: 8px;'>{act}</li>" for act in custom_actions])}
                    </ul>
                    <p style='font-size: 11px; color: #64748b; margin-top: 20px;'>
                        Auto-generated by KMP Centralised Security Data Management System.
                    </p>
                </div>
                """
                message = MessageSchema(
                    subject=f"Weekly Tactical Briefing: {station}",
                    recipients=[user.email],
                    body=html_body,
                    subtype="html"
                )
                
                try:
                    await fm.send_message(message)
                except Exception as mail_err:
                    print(f"Failed to dispatch to {user.email}: {mail_err}")

        asyncio.run(process_and_send_emails())

    except Exception as e:
        print(f"Dynamic scheduler error: {e}")
    finally:
        db.close()

scheduler = BackgroundScheduler()
scheduler.add_job(run_weekly_tactical_briefing_job, 'cron', day_of_week='mon', hour=6, minute=0)

@app.on_event("startup")
def start_scheduler():
    if not scheduler.running:
        scheduler.start()

@app.on_event("shutdown")
def shutdown_scheduler():
    scheduler.shutdown()

# ====================================================================
# SECURE ENCRYPTED ZIP EXPORTS (AUDIT LOGS & HR LEDGER)
# ====================================================================

@app.get("/api/v1/audit-logs/export")
def export_audit_logs_excel(db: Session = Depends(get_db), current_user: models.Users = Depends(get_current_user)):
    user_role = str(current_user.role).strip().upper() if current_user.role else ""
    if "ADMIN" not in user_role and "RPC" not in user_role:
        raise HTTPException(status_code=403, detail="Clearance Denied: Admin privileges required.")
        
    try:
        AuditModel = getattr(models, 'Audit_Logs', getattr(models, 'AuditLogs', None))
        logs = db.query(AuditModel).order_by(AuditModel.id.desc()).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Command Audit Logs"
        ws.append(["ID", "Event Type", "Target User", "Status", "Details", "Created At", "User FNUM"])

        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center")

        for log in logs:
            details_clean = clean_html_for_export(getattr(log, 'details', ''))
            ws.append([log.id, getattr(log, 'event_type', ''), getattr(log, 'target_user', ''), getattr(log, 'status', ''), details_clean, str(getattr(log, 'created_at', '')), getattr(log, 'user_fnum', '')])

        for col in ws.columns:
            col_letter = col[0].column_letter
            max_len = max([len(str(cell.value or '')) for cell in col], default=0)
            if col[0].value == "Details":
                ws.column_dimensions[col_letter].width = 50
                for cell in col:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        
        zip_stream = io.BytesIO()
        eat_time = datetime.now(pytz.timezone("Africa/Nairobi")).replace(tzinfo=None)
        fnum_clean = str(current_user.fnum).replace('/', '_').upper()
        
        with pyzipper.AESZipFile(zip_stream, 'w', compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES) as zf:
            zf.setpassword(str(current_user.fnum).strip().encode('utf-8'))
            zf.writestr(f"{fnum_clean}_Audit_Logs_{eat_time.strftime('%Y%m%d')}.xlsx", excel_stream.getvalue())

        zip_stream.seek(0)
        return StreamingResponse(
            zip_stream, media_type="application/zip",
            headers={'Content-Disposition': f'attachment; filename="SECURE_AUDIT_LOGS_{eat_time.strftime("%y%m%d")}.zip"', 'Access-Control-Expose-Headers': 'Content-Disposition'}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export audit logs: {str(e)}")

@app.get("/api/v1/hr/export-ledger")
def export_hr_ledger(db: Session = Depends(get_db), current_user: models.Users = Depends(require_export_privilege)):
    try:
        is_global = current_user.role in ['SUPER_ADMIN', 'ADMIN', 'RPC'] or str(current_user.region).upper() in ['KMP HEADQUARTERS', 'POLICE HEADQUARTERS']
        NomModel = getattr(models, 'Nominal_Roll', getattr(models, 'NominalRoll', None))
        EstModel = getattr(models, 'Establishments', getattr(models, 'establishments', None))
        ArcModel = getattr(models, 'NominalRollArchive', getattr(models, 'Nominal_Roll_Archive', None))        

        def get_orm_data(ModelClass, columns):
            if not ModelClass: return []
            try:
                query = db.query(ModelClass)
                if not is_global and hasattr(ModelClass, 'region'): query = query.filter(ModelClass.region == current_user.region)
                
                rows = []
                for r in query.all():
                    row_data = []
                    for col in columns:
                        val = getattr(r, col, '')
                        if isinstance(val, datetime):
                            val = val.strftime("%Y-%m-%d %H:%M")
                        elif isinstance(val, str) and col in ['narrative', 'comment', 'message', 'details', 'archive_reason']:
                            val = clean_html_for_export(val)
                        row_data.append(str(val) if val is not None else '')
                    rows.append(row_data)
                return rows
            except: return []

        nr_records = get_orm_data(NomModel, ['f_num', 'name', 'rank', 'sex', 'region', 'station', 'position', 'status'])
        est_records = get_orm_data(EstModel, ['region', 'division', 'station', 'personnel_in_station', 'sub_station', 'personnel_in_sub_station', 'post', 'personnel_in_post', 'booths', 'personnel_in_booth', 'installed_by', 'location', 'status', 'comment', 'last_updated_by'])
        arc_records = get_orm_data(ArcModel, ['fnum', 'name', 'rank', 'sex', 'region', 'station', 'position', 'status', 'archive_reason', 'archive_date'])

        wb = openpyxl.Workbook()
        wb.remove(wb.active) 
        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        def add_sheet(title, headers, data):
            ws = wb.create_sheet(title=title)
            ws.append(["SN"] + headers)
            for cell in ws[1]:
                cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center")
            for idx, row in enumerate(data, 1): ws.append([idx] + row)
            for col in ws.columns:
                max_len = max([len(str(cell.value or '')) for cell in col], default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 65)

        add_sheet("Nominal Roll", ["Force Number", "Name", "Rank", "Sex", "Region", "Station", "Position", "Status"], nr_records)
        add_sheet("establishments", ["Region", "Division", "Station", "Personnel (Station)", "Sub-Station", "Personnel (Sub-Stn)", "Post", "Personnel (Post)", "Booths", "Personnel (Booth)", "Installed By", "Location", "Status", "Comment", "Last Updated By"], est_records)
        add_sheet("Archived Personnel", ["Force Number", "Name", "Rank", "Sex", "Region", "Station", "Position", "Status", "Reason", "Archived On"], arc_records)

        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        
        zip_stream = io.BytesIO()
        eat_time = datetime.now(pytz.timezone("Africa/Nairobi")).replace(tzinfo=None)
        fnum_clean = str(current_user.fnum).replace('/', '_').upper()
        
        with pyzipper.AESZipFile(zip_stream, 'w', compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES) as zf:
            zf.setpassword(str(current_user.fnum).strip().encode('utf-8'))
            zf.writestr(f"{fnum_clean}_HR_Ledger_{eat_time.strftime('%Y%m%d')}.xlsx", excel_stream.getvalue())

        zip_stream.seek(0)
        return StreamingResponse(
            zip_stream, media_type="application/zip",
            headers={'Content-Disposition': f'attachment; filename="SECURE_HR_LEDGER_{eat_time.strftime("%y%m%d")}.zip"', 'Access-Control-Expose-Headers': 'Content-Disposition'}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"HR export failed: {str(e)}")

@app.get("/api/v1/reports/export")
def export_master_database(timeframe: str = "all", scope: Optional[str] = None, value: Optional[str] = None, db: Session = Depends(get_db), current_user: models.Users = Depends(require_export_privilege)):
    try:
        is_global = current_user.role in ['SUPER_ADMIN', 'ADMIN', 'RPC'] or str(current_user.region).upper() in ['KMP HEADQUARTERS', 'POLICE HEADQUARTERS']
        
        CrimeModel = getattr(models, 'Crime_Reports', getattr(models, 'CrimeReports', getattr(models, 'Reports', None)))
        StatsModel = getattr(models, 'Operational_Statistics', getattr(models, 'OperationalStatistics', getattr(models, 'Stats', None)))
        StoryModel = getattr(models, 'Success_Stories', getattr(models, 'SuccessStories', getattr(models, 'Stories', None)))
        NomModel = getattr(models, 'Nominal_Roll', getattr(models, 'NominalRoll', None))
        EstModel = getattr(models, 'Establishments', getattr(models, 'establishments', None))
        DocsModel = getattr(models, 'DocumentArchive', getattr(models, 'Document_Archive', getattr(models, 'document_archive', None)))
        ActivityModel = getattr(models, 'Activity_Logs', getattr(models, 'ActivityLogs', None))
        AIModel = getattr(models, 'AI_Command_Logs', getattr(models, 'AICommandLogs', None))
        ArcModel = getattr(models, 'NominalRollArchive', getattr(models, 'Nominal_Roll_Archive', None))

        def get_full_dataframe(ModelClass):
            if not ModelClass: 
                return pd.DataFrame()
            try:
                query = db.query(ModelClass)
                if not is_global and hasattr(ModelClass, 'region'): 
                    query = query.filter(ModelClass.region == current_user.region)
                
                records = query.all()
                if not records:
                    return pd.DataFrame()
                
                data = []
                for r in records:
                    row_dict = {}
                    for col in r.__table__.columns.keys():
                        val = getattr(r, col, '')
                        if isinstance(val, datetime):
                            val = val.strftime("%Y-%m-%d %H:%M")
                        elif isinstance(val, str) and col in ['narrative', 'comment', 'message', 'details', 'archive_reason']:
                            val = clean_html_for_export(val)
                        row_dict[col] = val if val is not None else ''
                    data.append(row_dict)
                return pd.DataFrame(data)
            except Exception as ex:
                print(f"DataFrame fetch error for {ModelClass}: {ex}")
                return pd.DataFrame()

        # Fetch full dataframes containing all NeonDB columns
        df_crime = get_full_dataframe(CrimeModel)
        df_stats = get_full_dataframe(StatsModel)
        df_stories = get_full_dataframe(StoryModel)
        df_users = get_full_dataframe(NomModel)
        df_est = get_full_dataframe(EstModel)
        df_docs = get_full_dataframe(DocsModel)
        df_arc = get_full_dataframe(ArcModel)

        # Compile AI & Activity Command logs
        ai_rows = []
        if AIModel:
            try:
                for r in db.query(AIModel).all():
                    v_time = r.created_at.strftime("%Y-%m-%d %H:%M") if hasattr(r.created_at, 'strftime') else str(getattr(r, 'created_at', ''))
                    ai_rows.append({"Interaction Type": "AI_PROMPT_EXECUTION", "Details": clean_html_for_export(f"Prompt: {r.prompt} | Response: {r.response}"), "Officer FNUM": getattr(r, 'fnum', ''), "Timestamp": v_time})
            except Exception: 
                pass
                
        if ActivityModel:
            try:
                ai_act = db.query(ActivityModel).filter(or_(ActivityModel.module.ilike('%AI%'), ActivityModel.module.ilike('%ai_console%'))).all()
                for r in ai_act:
                    v_time = r.created_at.strftime("%Y-%m-%d %H:%M") if hasattr(r.created_at, 'strftime') else str(getattr(r, 'created_at', ''))
                    ai_rows.append({"Interaction Type": "PAGE_ACCESS_LOG", "Details": clean_html_for_export(getattr(r, 'details', 'AI Console Access')), "Officer FNUM": getattr(r, 'fnum', ''), "Timestamp": v_time})
            except Exception: 
                pass
        df_ai = pd.DataFrame(ai_rows)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal="center", vertical="center")

        def write_dual_sheets(df, print_title, full_title, print_cols):
            if df.empty:
                return
            
            # 1. Print / Summarized Copy Sheet
            ws_print = wb.create_sheet(title=print_title)
            available_print_cols = [c for c in print_cols if c in df.columns]
            df_print_subset = df[available_print_cols].copy()
            
            ws_print.append(["SN"] + list(df_print_subset.columns))
            for cell in ws_print[1]:
                cell.fill = header_fill; cell.font = header_font; cell.alignment = header_align
            
            for idx, row in enumerate(df_print_subset.values, 1):
                ws_print.append([idx] + list(row))
                
            for col in ws_print.columns:
                max_len = max([len(str(cell.value or '')) for cell in col], default=0)
                ws_print.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

            # 2. Full NeonDB Columns Copy Sheet
            ws_full = wb.create_sheet(title=full_title)
            ws_full.append(["SN"] + list(df.columns))
            for cell in ws_full[1]:
                cell.fill = header_fill; cell.font = header_font; cell.alignment = header_align
                
            for idx, row in enumerate(df.values, 1):
                ws_full.append([idx] + list(row))
                
            for col in ws_full.columns:
                max_len = max([len(str(cell.value or '')) for cell in col], default=0)
                ws_full.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)

        # Generate dual sheets for all domains
        write_dual_sheets(df_crime, "Crime Registry (Print)", "Crime Registry", ['sd_ref', 'region', 'station', 'date', 'time', 'offence', 'status', 'suspects', 'last_updated_by'])
        write_dual_sheets(df_stats, "OPS Statistics (Print)", "OPS Statistics", ['date', 'region', 'station', 'arrested', 'given_bond', 'cautioned', 'pending_court', 'taken_to_court', 'released', 'remanded', 'convicted'])
        write_dual_sheets(df_stories, "Success Stories (Print)", "Success Stories", ['date', 'time', 'region', 'station', 'status', 'narrative'])
        write_dual_sheets(df_users, "Establishments (Print)", "Nominal Roll", ['f_num', 'name', 'rank', 'sex', 'region', 'station', 'position', 'status'])
        write_dual_sheets(df_arc, "Archived Personnel (Print)", "Archived Personnel", ['fnum', 'name', 'rank', 'sex', 'region', 'station', 'position', 'status', 'archive_reason', 'archive_date'])
        write_dual_sheets(df_est, "Establishments Print Copy", "Establishments", ['region', 'division', 'station', 'personnel_in_station', 'sub_station', 'personnel_in_sub_station', 'post', 'personnel_in_post', 'booths', 'personnel_in_booth'])
        write_dual_sheets(df_docs, "Tripartite Reports (Print)", "Tripartite Reports", ['file_name', 'doc_type', 'file_size', 'region', 'station', 'uploaded_by', 'upload_date'])
        write_dual_sheets(df_ai, "AI Command (Print)", "AI Command", ['Interaction Type', 'Details', 'Officer FNUM', 'Timestamp'])

        eat_tz = pytz.timezone("Africa/Nairobi")
        eat_time = datetime.now(eat_tz).replace(tzinfo=None)
        
        officer_fnum = (current_user.fnum or "HQ-UNKNOWN").strip().upper()
        stamp_id = f"KMP-STAMP-{officer_fnum}-{eat_time.strftime('%Y%m%d%H%M%S')}"
        encoded_token = base64.b64encode(json.dumps({"f": officer_fnum, "s": stamp_id}).encode('utf-8')).decode('utf-8')
        
        wb.properties.keywords = f"KMP_AUDIT;{encoded_token}"
        wb.properties.category = "RESTRICTED / FORENSIC POLICE RECORD"

        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        
        zip_stream = io.BytesIO()
        zip_password = str(current_user.fnum).strip().encode('utf-8')
        fnum_clean = str(current_user.fnum).replace('/', '_').upper()

        with pyzipper.AESZipFile(zip_stream, 'w', compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES) as zf:
            zf.setpassword(zip_password)
            excel_filename = f"{fnum_clean}_Master_Database_{eat_time.strftime('%Y%m%d')}.xlsx"
            zf.writestr(excel_filename, excel_stream.getvalue())

        zip_stream.seek(0)
        zip_filename = f"SECURE_MASTER_DB_{eat_time.strftime('%Y%m%d')}.zip"
        
        return StreamingResponse(
            zip_stream, media_type="application/zip",
            headers={'Content-Disposition': f'attachment; filename="{zip_filename}"', 'Access-Control-Expose-Headers': 'Content-Disposition'}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Master export failed: {str(e)}")

@app.get("/api/v1/export/establishments")
def export_establishments_summary(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    try:
        est_model = getattr(models, 'Establishments', getattr(models, 'establishments', None))
        est = db.query(est_model).all() if est_model else []
        df_est = sanitize_df_for_excel(pd.DataFrame([serialize_model_row(e) for e in est]))
        
        output = io.BytesIO()
        df_est.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=HR_Establishments_Summary.xlsx"}
        )
    except Exception as e:
        print(f"HR export error: {e}")
        raise HTTPException(status_code=500, detail=f"HR export failed: {str(e)}")

if __name__ == "__main__":
    uvicorn.run("api_backend:app", host="0.0.0.0", port=8000, reload=True)